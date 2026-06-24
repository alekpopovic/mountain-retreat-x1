from pathlib import Path

from openpyxl import load_workbook
from typer.testing import CliRunner

from mountain_retreat_x1.cli import app
from mountain_retreat_x1.exporters import BOM_FILENAME
from mountain_retreat_x1.exporters.excel import ITEM_HEADERS, WORKBOOK_SHEETS

runner = CliRunner()


def test_generate_excel_bom_creates_workbook_with_expected_sheets_and_headers(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"

    result = runner.invoke(
        app,
        ["generate", "excel", "--bom", "--output-dir", str(output_dir)],
    )

    assert result.exit_code == 0
    assert "Excel BOM generation completed" in result.output
    workbook_path = output_dir / "excel" / BOM_FILENAME
    assert workbook_path.exists()

    workbook = load_workbook(workbook_path, data_only=False)
    assert workbook.sheetnames == list(WORKBOOK_SHEETS)
    for sheet_name in WORKBOOK_SHEETS:
        sheet = workbook[sheet_name]
        if sheet_name in ("Summary", "Assumptions"):
            continue
        headers = [cell.value for cell in sheet[1]]
        assert headers == list(ITEM_HEADERS)
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref is not None


def test_generate_excel_bom_contains_formulas_and_assumptions(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"

    result = runner.invoke(
        app,
        ["generate", "excel", "--bom", "--output-dir", str(output_dir)],
    )

    assert result.exit_code == 0
    workbook = load_workbook(output_dir / "excel" / BOM_FILENAME, data_only=False)
    foundations = workbook["Concrete_and_Foundations"]
    assert foundations["I2"].value == "=G2*(1+H2/100)"
    assert foundations["M2"].value == "=I2*J2"
    assert foundations["N2"].value == "=I2*K2"
    assert foundations["O2"].value == "=I2*L2"
    assert "Supplier placeholder:" in str(foundations["R2"].value)

    summary = workbook["Summary"]
    assert str(summary["B2"].value).startswith("=SUMIF(")
    assert "TOTAL" in [cell.value for cell in summary["A"]]

    assumptions = workbook["Assumptions"]
    assumption_names = [cell.value for cell in assumptions["A"]]
    assumption_values = [cell.value for cell in assumptions["B"]]
    assert "Status" in assumption_names
    assert "PRELIMINARY" in assumption_values
    assert "Disclaimer" in assumption_names
