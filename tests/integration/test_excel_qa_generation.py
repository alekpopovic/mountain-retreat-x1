from pathlib import Path

from openpyxl import load_workbook
from typer.testing import CliRunner

from mountain_retreat_x1.cli import app
from mountain_retreat_x1.exporters import QA_FILENAME
from mountain_retreat_x1.exporters.excel import QA_HEADERS, QA_STATUS_VALUES, QA_WORKBOOK_SHEETS

runner = CliRunner()


def test_generate_excel_qa_creates_expected_sheets_and_headers(tmp_path: Path) -> None:
    output_dir = tmp_path / "output"

    result = runner.invoke(
        app,
        ["generate", "excel", "--qa", "--output-dir", str(output_dir)],
    )

    assert result.exit_code == 0
    assert "Excel generation completed" in result.output
    workbook_path = output_dir / "excel" / QA_FILENAME
    assert workbook_path.exists()

    workbook = load_workbook(workbook_path, data_only=False)
    assert workbook.sheetnames == list(QA_WORKBOOK_SHEETS)
    for sheet_name in QA_WORKBOOK_SHEETS:
        sheet = workbook[sheet_name]
        assert [cell.value for cell in sheet[1]] == list(QA_HEADERS)
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref is not None


def test_generate_excel_qa_has_large_mode_rows_and_status_validation(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"

    result = runner.invoke(
        app,
        ["generate", "excel", "--qa", "--large", "--output-dir", str(output_dir)],
    )

    assert result.exit_code == 0
    workbook = load_workbook(output_dir / "excel" / QA_FILENAME, data_only=False)
    total_rows = sum(workbook[sheet_name].max_row - 1 for sheet_name in QA_WORKBOOK_SHEETS)
    assert total_rows >= 1000

    site_sheet = workbook["Site_Preparation"]
    validation = next(iter(site_sheet.data_validations.dataValidation))
    assert validation.type == "list"
    for status in QA_STATUS_VALUES:
        assert status in validation.formula1
    assert "I2:I53" in str(validation.sqref)
    assert site_sheet["I2"].value == "Not Started"

    terrace_sheet = workbook["Terrace"]
    inspection_items = [cell.value for cell in terrace_sheet["D"]]
    assert any("Membrane" in str(item) for item in inspection_items)
    assert any("Guard" in str(item) for item in inspection_items)


def test_generate_excel_qa_large_mode_has_more_rows_than_normal(
    tmp_path: Path,
) -> None:
    normal_dir = tmp_path / "normal"
    large_dir = tmp_path / "large"

    normal_result = runner.invoke(
        app,
        ["generate", "excel", "--qa", "--output-dir", str(normal_dir)],
    )
    large_result = runner.invoke(
        app,
        ["generate", "excel", "--qa", "--large", "--output-dir", str(large_dir)],
    )

    assert normal_result.exit_code == 0
    assert large_result.exit_code == 0
    normal_workbook = load_workbook(normal_dir / "excel" / QA_FILENAME, data_only=False)
    large_workbook = load_workbook(large_dir / "excel" / QA_FILENAME, data_only=False)
    normal_rows = sum(
        normal_workbook[sheet_name].max_row - 1 for sheet_name in QA_WORKBOOK_SHEETS
    )
    large_rows = sum(
        large_workbook[sheet_name].max_row - 1 for sheet_name in QA_WORKBOOK_SHEETS
    )
    assert normal_rows < large_rows
    assert large_rows >= 1000
