from pathlib import Path

from openpyxl import load_workbook
from typer.testing import CliRunner

from mountain_retreat_x1.cli import app
from mountain_retreat_x1.exporters import GANTT_FILENAME
from mountain_retreat_x1.exporters.excel import GANTT_HEADERS, GANTT_WORKBOOK_SHEETS

runner = CliRunner()


def test_generate_excel_gantt_creates_workbook_with_expected_sheets(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"

    result = runner.invoke(
        app,
        ["generate", "excel", "--gantt", "--output-dir", str(output_dir)],
    )

    assert result.exit_code == 0
    assert "Excel generation completed" in result.output
    workbook_path = output_dir / "excel" / GANTT_FILENAME
    assert workbook_path.exists()

    workbook = load_workbook(workbook_path, data_only=False)
    assert workbook.sheetnames == list(GANTT_WORKBOOK_SHEETS)
    assert [cell.value for cell in workbook["Gantt"][1]] == list(GANTT_HEADERS)


def test_generate_excel_gantt_has_52_week_columns_and_timeline_fills(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"

    result = runner.invoke(
        app,
        ["generate", "excel", "--gantt", "--output-dir", str(output_dir)],
    )

    assert result.exit_code == 0
    workbook = load_workbook(output_dir / "excel" / GANTT_FILENAME, data_only=False)
    gantt = workbook["Gantt"]
    week_headers = [cell.value for cell in gantt[1][12:64]]
    assert week_headers == [f"Week {week}" for week in range(1, 53)]
    assert gantt.max_column == 64
    assert gantt["M2"].value == "■"
    assert gantt["M2"].fill.fgColor.rgb in {"0070AD47", "70AD47"}
    assert gantt.freeze_panes == "M2"
    assert gantt.auto_filter.ref is not None

    milestones = workbook["Milestones"]
    milestone_names = [cell.value for cell in milestones["A"]]
    assert "Weather-tight shell" in milestone_names

    dependencies = workbook["Dependencies"]
    warnings = [cell.value for cell in dependencies["D"]]
    assert any("predecessors" in str(warning) for warning in warnings)

    assumptions = workbook["Assumptions"]
    assumption_values = [cell.value for cell in assumptions["B"]]
    assert "12 months / 52 weeks" in assumption_values
    assert any("Snow, frost, rain, wind" in str(value) for value in assumption_values)
