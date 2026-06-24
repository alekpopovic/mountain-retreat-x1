from pathlib import Path

from openpyxl import load_workbook
from typer.testing import CliRunner

from mountain_retreat_x1.cli import app
from mountain_retreat_x1.exporters import MAINTENANCE_FILENAME
from mountain_retreat_x1.exporters.excel import (
    CALENDAR_HEADERS,
    MAINTENANCE_WORKBOOK_SHEETS,
)

runner = CliRunner()


def test_generate_excel_maintenance_creates_calendar_workbook(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"

    result = runner.invoke(
        app,
        ["generate", "excel", "--maintenance", "--output-dir", str(output_dir)],
    )

    assert result.exit_code == 0
    assert "Excel generation completed" in result.output
    workbook_path = output_dir / "excel" / MAINTENANCE_FILENAME
    assert workbook_path.exists()

    workbook = load_workbook(workbook_path, data_only=False)
    assert workbook.sheetnames == list(MAINTENANCE_WORKBOOK_SHEETS)
    assert [cell.value for cell in workbook["Calendar"][1]] == list(CALENDAR_HEADERS)
    assert workbook["Calendar"].freeze_panes == "A2"
    assert workbook["Calendar"].auto_filter.ref is not None


def test_generate_excel_maintenance_contains_30_year_planning_rows(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"

    result = runner.invoke(
        app,
        ["generate", "excel", "--maintenance", "--output-dir", str(output_dir)],
    )

    assert result.exit_code == 0
    workbook = load_workbook(
        output_dir / "excel" / MAINTENANCE_FILENAME,
        data_only=False,
    )
    calendar_years = {cell.value for cell in workbook["Calendar"]["A"][1:]}
    assert {1, 5, 10, 20, 30}.issubset(calendar_years)

    year_plan = workbook["Year_1_to_30"]
    assert year_plan.max_row == 31
    focus_values = [cell.value for cell in year_plan["B"]]
    assert "5-year maintenance and sealant/coating review" in focus_values
    assert "10-year maintenance and medium renewal review" in focus_values
    assert "Major system renewal study" in focus_values
    assert "30-year renewal planning" in focus_values

    catalog_tasks = [cell.value for cell in workbook["Task_Catalog"]["B"]]
    assert any("Solar PV" in str(task) or "PV production" in str(task) for task in catalog_tasks)
    assert any("Battery" in str(task) or "BMS" in str(task) for task in catalog_tasks)
