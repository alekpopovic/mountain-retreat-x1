from pathlib import Path

import pytest
from openpyxl import load_workbook
from typer.testing import CliRunner

from mountain_retreat_x1.cli import app
from mountain_retreat_x1.exporters import BOM_FILENAME, COST_FILENAME

runner = CliRunner()


@pytest.mark.parametrize(
    ("variant", "name", "material_code", "cost_code"),
    [
        ("standard_hybrid", "Standard timber hybrid", "VAR-STD-MAT-001", "VAR-STD-COST-001"),
        ("premium_clt", "Premium CLT/glulam", "VAR-CLT-MAT-001", "VAR-CLT-COST-001"),
        ("masonry_hybrid", "Masonry hybrid", "VAR-MAS-MAT-001", "VAR-MAS-COST-001"),
    ],
)
def test_generate_all_supports_construction_variants(
    tmp_path: Path,
    variant: str,
    name: str,
    material_code: str,
    cost_code: str,
) -> None:
    output_dir = tmp_path / variant

    result = runner.invoke(
        app,
        [
            "generate",
            "all",
            "--variant",
            variant,
            "--lang",
            "en",
            "--output",
            str(output_dir),
        ],
    )

    assert result.exit_code == 0

    charter = (output_dir / "markdown" / "01_project_charter.md").read_text(encoding="utf-8")
    structural = (output_dir / "markdown" / "03_structural_concept.md").read_text(
        encoding="utf-8"
    )
    self_build = (output_dir / "markdown" / "13_self_build_guide.md").read_text(
        encoding="utf-8"
    )
    construction_management = (
        output_dir / "markdown" / "11_construction_management.md"
    ).read_text(encoding="utf-8")

    assert name in charter
    assert f"`{variant}`" in charter
    assert name in structural
    assert f"config/variants/{variant}.yaml" in structural
    assert "No variant in this document is structurally approved" in structural
    assert name in self_build
    assert "Active Variant Self-Build Warnings" in self_build
    assert name in construction_management

    bom = load_workbook(output_dir / "excel" / BOM_FILENAME, data_only=False)
    bom_values = [
        str(cell.value)
        for sheet in bom.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert material_code in bom_values
    assert name in bom_values

    cost = load_workbook(output_dir / "excel" / COST_FILENAME, data_only=False)
    cost_values = [
        str(cell.value)
        for sheet in cost.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert cost_code in cost_values
    assert name in cost_values
