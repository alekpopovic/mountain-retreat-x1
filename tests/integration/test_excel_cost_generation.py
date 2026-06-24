from pathlib import Path

from openpyxl import load_workbook
from typer.testing import CliRunner

from mountain_retreat_x1.cli import app
from mountain_retreat_x1.exporters import COST_FILENAME
from mountain_retreat_x1.exporters.excel import COST_WORKBOOK_SHEETS

runner = CliRunner()


def test_generate_excel_cost_creates_workbook_with_expected_sheets(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"

    result = runner.invoke(
        app,
        ["generate", "excel", "--cost", "--output-dir", str(output_dir)],
    )

    assert result.exit_code == 0
    assert "Excel generation completed" in result.output
    workbook_path = output_dir / "excel" / COST_FILENAME
    assert workbook_path.exists()

    workbook = load_workbook(workbook_path, data_only=False)
    assert workbook.sheetnames == list(COST_WORKBOOK_SHEETS)
    for sheet_name in COST_WORKBOOK_SHEETS:
        sheet = workbook[sheet_name]
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref is not None


def test_generate_excel_cost_contains_formulas_summary_totals_and_notes(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"

    result = runner.invoke(
        app,
        ["generate", "excel", "--cost", "--output-dir", str(output_dir)],
    )

    assert result.exit_code == 0
    workbook = load_workbook(output_dir / "excel" / COST_FILENAME, data_only=False)

    materials = workbook["Materials"]
    assert materials["H2"].value == "=E2*G2"

    contingency = workbook["Contingency"]
    assert contingency["C2"].value == "=Materials!H2+Labor!H2+Machinery!H2"
    assert contingency["F2"].value == "=C2*10%"
    assert contingency["G2"].value == "=C2*15%"
    assert contingency["H2"].value == "=C2*20%"
    assert contingency["I2"].value == "=C2+E2"

    summary = workbook["Executive_Summary"]
    summary_values = {summary[f"A{row}"].value: summary[f"B{row}"].value for row in range(2, 19)}
    assert summary_values["Material subtotal"] == "=Materials!H5"
    assert summary_values["Labor subtotal"] == "=Labor!H5"
    assert summary_values["Machinery subtotal"] == "=Machinery!H5"
    assert summary_values["Subtotal before contingency"] == "=SUM(B2:B4)"
    assert summary_values["Contingency 10%"] == "=B5*10%"
    assert summary_values["Contingency 15%"] == "=B5*15%"
    assert summary_values["Contingency 20%"] == "=B5*20%"
    assert summary_values["Total with contingency"] == "=B5+B8"
    assert summary_values["Cost per gross m2"] == "=B9/Assumptions!$B$11"
    assert summary_values["Cost per net m2"] == "=B9/Assumptions!$B$12"
    assert summary_values["Optional off-grid add-on"] == "=Scenario_OffGrid_AddOn!F7"
    assert summary_values["Economy scenario"] == "=Scenario_Economy!G5"

    off_grid = workbook["Scenario_OffGrid_AddOn"]
    assert off_grid["F2"].value == "=C2*E2"
    assert off_grid["F7"].value == "=SUM(F2:F6)"

    cash_flow = workbook["Cash_Flow"]
    assert cash_flow["C2"].value == "=B2/SUM($B$2:$B$4)"

    assumptions = workbook["Assumptions"]
    assumption_names = [cell.value for cell in assumptions["A"]]
    assumption_values = [cell.value for cell in assumptions["B"]]
    assert "last_updated" in assumption_names
    assert "2026-06-24" in assumption_values

    notes = [cell.value for cell in workbook["Review_Notes"]["A"]]
    assert "All prices are placeholders and static planning assumptions." in notes
    assert any("Contractor quotes are required" in str(note) for note in notes)
