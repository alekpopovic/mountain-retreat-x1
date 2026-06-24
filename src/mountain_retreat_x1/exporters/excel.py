"""Excel workbook exporters for preliminary planning documents."""

from collections.abc import Iterable
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from mountain_retreat_x1.config.loader import MountainRetreatConfig
from mountain_retreat_x1.models import CostItem, MaterialItem

BOM_FILENAME = "Mountain_Retreat_X1_BOM.xlsx"
COST_FILENAME = "Mountain_Retreat_X1_Cost_Estimate.xlsx"
GANTT_FILENAME = "Mountain_Retreat_X1_Gantt_Schedule.xlsx"
QA_FILENAME = "Mountain_Retreat_X1_QA_QC_Checklists.xlsx"
MAINTENANCE_FILENAME = "Mountain_Retreat_X1_Maintenance_Calendar.xlsx"
BOM_OUTPUT_DIR = "excel"
COST_OUTPUT_DIR = "excel"
GANTT_OUTPUT_DIR = "excel"
QA_OUTPUT_DIR = "excel"
MAINTENANCE_OUTPUT_DIR = "excel"
DETERMINISTIC_WORKBOOK_TIMESTAMP = datetime(2026, 6, 24, tzinfo=UTC)

ITEM_HEADERS = (
    "Item Code",
    "Category",
    "Subcategory",
    "Description",
    "Specification",
    "Unit",
    "Base Quantity",
    "Waste %",
    "Quantity With Waste",
    "Unit Cost Low",
    "Unit Cost Standard",
    "Unit Cost Premium",
    "Subtotal Low",
    "Subtotal Standard",
    "Subtotal Premium",
    "Formula Note",
    "Assumption Ref",
    "Notes",
)

ITEM_SHEETS = (
    "Concrete_and_Foundations",
    "Rebar_and_Steel",
    "Timber_and_Glulam",
    "CLT_Optional",
    "Masonry_Optional",
    "Roof",
    "Facade",
    "Windows_and_Doors",
    "Insulation_and_Membranes",
    "Terrace",
    "Electrical",
    "Plumbing",
    "HVAC",
    "Smart_Home",
    "Off_Grid",
    "Interior_Finishes",
    "Tools",
    "Machinery",
    "Waste_Factors",
)

WORKBOOK_SHEETS = ("Summary", *ITEM_SHEETS, "Assumptions")
COST_WORKBOOK_SHEETS = (
    "Executive_Summary",
    "Cost_By_Phase",
    "Cost_By_Category",
    "Materials",
    "Labor",
    "Machinery",
    "Contingency",
    "Cash_Flow",
    "Scenario_Economy",
    "Scenario_Standard",
    "Scenario_Premium",
    "Scenario_OffGrid_AddOn",
    "Assumptions",
    "Review_Notes",
)
GANTT_WORKBOOK_SHEETS = (
    "Gantt",
    "Phase_Details",
    "Milestones",
    "Dependencies",
    "Risks",
    "Inspections",
    "Cashflow_By_Week",
    "Assumptions",
)
QA_WORKBOOK_SHEETS = (
    "Site_Preparation",
    "Earthworks",
    "Drainage",
    "Foundations",
    "Rebar",
    "Concrete",
    "Timber_Structure",
    "Roof",
    "Windows_Doors",
    "Facade",
    "Electrical",
    "Plumbing",
    "HVAC",
    "Insulation",
    "Interior",
    "Terrace",
    "Smart_Home",
    "Solar_OffGrid",
    "Final_Inspection",
    "Photo_Log",
    "Document_Register",
)
QA_ASSUMPTIONS_SHEET = "Assumptions"

COST_ITEM_HEADERS = (
    "Cost Code",
    "Phase",
    "Category",
    "Description",
    "Quantity",
    "Unit",
    "Unit Cost",
    "Subtotal",
    "Notes",
)

SCENARIO_HEADERS = (
    "Cost Code",
    "Description",
    "Base Subtotal Before Contingency",
    "Scenario Factor",
    "Scenario Subtotal",
    "Contingency %",
    "Total With Contingency",
    "Notes",
)

EUR_FORMAT = '#,##0.00 "EUR"'
WEEK_HEADERS = tuple(f"Week {week}" for week in range(1, 53))
GANTT_HEADERS = (
    "WBS",
    "Phase",
    "Task",
    "Duration Days",
    "Start Week",
    "End Week",
    "Dependencies",
    "Responsible Party",
    "Risk Level",
    "Inspection Required",
    "Weather Sensitive",
    "Notes",
    *WEEK_HEADERS,
)
QA_HEADERS = (
    "Checklist ID",
    "Phase",
    "Discipline",
    "Inspection Item",
    "Acceptance Criteria",
    "Responsible Person",
    "Required Photo",
    "Required Document",
    "Status",
    "Date",
    "Notes",
    "Reviewer",
)
QA_STATUS_VALUES = (
    "Not Started",
    "In Progress",
    "Passed",
    "Failed",
    "Needs Review",
    "Not Applicable",
)

MAINTENANCE_WORKBOOK_SHEETS = (
    "Calendar",
    "Task_Catalog",
    "Year_1_to_30",
    "Professional_Reviews",
    "Emergency_Plan",
    "Maintenance_Log",
    "Assumptions",
)

MAINTENANCE_HEADERS = (
    "Task",
    "Frequency",
    "Responsible Person",
    "Estimated Effort",
    "Required Tools",
    "Warning Signs",
    "Notes",
)

CALENDAR_HEADERS = (
    "Year",
    "Month/Season",
    "Section",
    *MAINTENANCE_HEADERS,
    "Professional Review Required",
    "Completed Date",
    "Photo/Document Ref",
    "Closeout Notes",
)


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
NOTICE_FONT = Font(bold=True, color="9C0006")
TIMELINE_FILL = PatternFill("solid", fgColor="70AD47")
MILESTONE_FILL = PatternFill("solid", fgColor="FFC000")
RISK_FILL = PatternFill("solid", fgColor="F4B084")


@dataclass(frozen=True)
class GanttTask:
    """Preliminary Gantt task row."""

    wbs: str
    phase: str
    task: str
    duration_days: int
    start_week: int
    end_week: int
    dependencies: str
    responsible_party: str
    risk_level: str
    inspection_required: str
    weather_sensitive: str
    notes: str


@dataclass(frozen=True)
class MaintenanceCalendarTask:
    """Maintenance calendar task row."""

    section: str
    task: str
    frequency: str
    responsible_person: str
    estimated_effort: str
    required_tools: str
    warning_signs: str
    notes: str
    month_season: str
    due_years: tuple[int, ...]
    professional_review_required: str


def _large_mode_materials(config: MountainRetreatConfig) -> list[MaterialItem]:
    items: list[MaterialItem] = []
    for index, room in enumerate(
        [*config.rooms_ground_floor.rooms, *config.rooms_gallery.rooms],
        start=1,
    ):
        items.extend(
            (
                MaterialItem(
                    code=f"LM-RM-{index:03d}-FLOOR",
                    category="Interior finishes",
                    subcategory="Room finish",
                    name=f"{room.name} floor finish allowance",
                    specification=room.floor_finish,
                    unit="m2",
                    base_quantity=room.area_m2,
                    waste_percent=8,
                    unit_price_low=18,
                    unit_price_standard=35,
                    unit_price_premium=65,
                    supplier_placeholder="Supplier TBD",
                    formula_note="Room area from YAML room.area_m2.",
                    assumption_ref=f"rooms.{room.code}.area_m2",
                    notes="Large-mode planning line item; verify finish build-up.",
                ),
                MaterialItem(
                    code=f"LM-RM-{index:03d}-WALL",
                    category="Interior finishes",
                    subcategory="Wall finish",
                    name=f"{room.name} wall finish allowance",
                    specification=room.wall_finish,
                    unit="m2",
                    base_quantity=2 * (room.length_m + room.width_m) * room.clear_height_m,
                    waste_percent=10,
                    unit_price_low=10,
                    unit_price_standard=22,
                    unit_price_premium=45,
                    supplier_placeholder="Supplier TBD",
                    formula_note="Perimeter x clear height from YAML room dimensions.",
                    assumption_ref=f"rooms.{room.code}.dimensions",
                    notes="Openings not deducted; preliminary only.",
                ),
                MaterialItem(
                    code=f"LM-RM-{index:03d}-CEIL",
                    category="Interior finishes",
                    subcategory="Ceiling finish",
                    name=f"{room.name} ceiling finish allowance",
                    specification=room.ceiling_finish,
                    unit="m2",
                    base_quantity=room.area_m2,
                    waste_percent=8,
                    unit_price_low=12,
                    unit_price_standard=25,
                    unit_price_premium=50,
                    supplier_placeholder="Supplier TBD",
                    formula_note="Room area from YAML room.area_m2.",
                    assumption_ref=f"rooms.{room.code}.area_m2",
                    notes="Large-mode planning line item; coordinate lighting and HVAC.",
                ),
            )
        )
    for index, zone in enumerate(config.terrace.zones, start=1):
        items.append(
            MaterialItem(
                code=f"LM-TER-{index:03d}",
                category="Terrace",
                subcategory="Zone finish",
                name=f"{zone.name} terrace finish allowance",
                specification=zone.surface_finish,
                unit="m2",
                base_quantity=zone.area_m2,
                waste_percent=10,
                unit_price_low=35,
                unit_price_standard=70,
                unit_price_premium=120,
                supplier_placeholder="Supplier TBD",
                formula_note="Terrace zone area from YAML terrace.zones.",
                assumption_ref=f"terrace.{zone.code}.area_m2",
                notes="Large-mode line item; waterproofing/drainage review required.",
            )
        )
    return items


def _all_materials(
    config: MountainRetreatConfig,
    *,
    large_mode: bool = False,
) -> list[MaterialItem]:
    materials = [
        item
        for item in [*config.materials_core.materials, *config.materials_mep.materials]
        if item.code != "MAT-CORE-002"
    ]
    materials.extend(config.variant.material_items)
    if large_mode:
        materials.extend(_large_mode_materials(config))
    return materials


def _sheet_for_material(item: MaterialItem) -> str:
    text = f"{item.category} {item.subcategory} {item.name}".lower()
    if any(marker in text for marker in ("foundation", "concrete", "slab")):
        return "Concrete_and_Foundations"
    if any(marker in text for marker in ("rebar", "steel", "reinforcement")):
        return "Rebar_and_Steel"
    if any(marker in text for marker in ("timber", "glulam", "lvl", "wood")):
        return "Timber_and_Glulam"
    if "clt" in text:
        return "CLT_Optional"
    if any(marker in text for marker in ("masonry", "block", "brick")):
        return "Masonry_Optional"
    if "roof" in text:
        return "Roof"
    if "facade" in text or "cladding" in text:
        return "Facade"
    if any(marker in text for marker in ("window", "door", "glazing")):
        return "Windows_and_Doors"
    if any(marker in text for marker in ("insulation", "membrane", "barrier")):
        return "Insulation_and_Membranes"
    if any(marker in text for marker in ("terrace", "decking", "outdoor")):
        return "Terrace"
    if any(marker in text for marker in ("smart home", "zigbee", "matter", "mqtt")):
        return "Smart_Home"
    if any(marker in text for marker in ("pv", "battery", "generator", "off-grid")):
        return "Off_Grid"
    if "electrical" in text:
        return "Electrical"
    if any(marker in text for marker in ("plumbing", "water", "wastewater", "tank")):
        return "Plumbing"
    if any(marker in text for marker in ("hvac", "heating", "mechanical", "heat pump")):
        return "HVAC"
    if any(marker in text for marker in ("finish", "interior", "floor", "wall", "ceiling")):
        return "Interior_Finishes"
    if "tool" in text:
        return "Tools"
    if any(marker in text for marker in ("machine", "machinery", "equipment")):
        return "Machinery"
    return "Interior_Finishes"


def _materials_by_sheet(materials: Iterable[MaterialItem]) -> dict[str, list[MaterialItem]]:
    grouped: dict[str, list[MaterialItem]] = {sheet: [] for sheet in ITEM_SHEETS}
    for item in materials:
        grouped[_sheet_for_material(item)].append(item)
    return grouped


def _style_header_row(sheet: Worksheet, row: int = 1) -> None:
    for cell in sheet[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _set_widths(sheet: Worksheet, widths: tuple[float, ...]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _configure_item_sheet(sheet: Worksheet) -> None:
    sheet.append(ITEM_HEADERS)
    _style_header_row(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(ITEM_HEADERS))}1"
    _set_widths(
        sheet,
        (
            16,
            16,
            22,
            32,
            46,
            12,
            14,
            10,
            18,
            14,
            18,
            18,
            16,
            20,
            18,
            42,
            26,
            48,
        ),
    )


def _append_material_row(sheet: Worksheet, item: MaterialItem) -> None:
    row = sheet.max_row + 1
    notes = f"{item.notes} Supplier placeholder: {item.supplier_placeholder}".strip()
    sheet.append(
        (
            item.code,
            item.category,
            item.subcategory,
            item.name,
            item.specification,
            item.unit,
            item.base_quantity,
            item.waste_percent,
            f"=G{row}*(1+H{row}/100)",
            item.unit_price_low,
            item.unit_price_standard,
            item.unit_price_premium,
            f"=I{row}*J{row}",
            f"=I{row}*K{row}",
            f"=I{row}*L{row}",
            item.formula_note,
            item.assumption_ref,
            notes,
        )
    )
    for cell in sheet[row]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in ("G", "H", "I", "J", "K", "L", "M", "N", "O"):
        sheet[f"{column}{row}"].number_format = "#,##0.00"


def _build_summary_sheet(
    sheet: Worksheet,
    materials: Iterable[MaterialItem],
) -> None:
    headers = (
        "Category",
        "Subtotal Low",
        "Subtotal Standard",
        "Subtotal Premium",
        "Source Sheets",
        "Notes",
    )
    sheet.append(headers)
    _style_header_row(sheet)
    categories = sorted({item.category for item in materials})
    source_sheets = ", ".join(ITEM_SHEETS)
    for row, category in enumerate(categories, start=2):
        low_formula = "+".join(
            f"SUMIF('{sheet_name}'!$B:$B,$A{row},'{sheet_name}'!$M:$M)"
            for sheet_name in ITEM_SHEETS
        )
        standard_formula = "+".join(
            f"SUMIF('{sheet_name}'!$B:$B,$A{row},'{sheet_name}'!$N:$N)"
            for sheet_name in ITEM_SHEETS
        )
        premium_formula = "+".join(
            f"SUMIF('{sheet_name}'!$B:$B,$A{row},'{sheet_name}'!$O:$O)"
            for sheet_name in ITEM_SHEETS
        )
        sheet.append(
            (
                category,
                f"={low_formula}",
                f"={standard_formula}",
                f"={premium_formula}",
                source_sheets,
                "Preliminary category subtotal from BOM item sheets.",
            )
        )
    total_row = sheet.max_row + 1
    sheet.append(
        (
            "TOTAL",
            f"=SUM(B2:B{total_row - 1})",
            f"=SUM(C2:C{total_row - 1})",
            f"=SUM(D2:D{total_row - 1})",
            "",
            "Preliminary total; not a procurement quote.",
        )
    )
    for cell in sheet[total_row]:
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = "#,##0.00"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:F{sheet.max_row}"
    _set_widths(sheet, (22, 18, 20, 18, 48, 42))


def _assumption_rows(config: MountainRetreatConfig) -> tuple[tuple[str, str], ...]:
    return (
        ("Project name", config.project.project_name),
        ("Project code", config.project.project_code),
        ("Version", config.project.version),
        ("Status", config.project.status),
        ("Disclaimer", config.project.disclaimer),
        ("Country", config.project.country),
        ("Currency", config.project.currency),
        ("Revision date", config.project.revision_date.isoformat()),
        ("Author", config.project.author),
        ("Professional review required by", ", ".join(config.project.review_required_by)),
        ("Gross area m2", f"{config.building.gross_area_m2:g}"),
        ("Net area m2", f"{config.building.net_area_m2:g}"),
        ("Terrace area m2", f"{config.terrace.terrace_area_m2:g}"),
        ("Construction variant", config.building.construction_variant),
        ("Active variant name", config.variant.name),
        ("Variant description", config.variant.description),
        ("Variant procurement complexity", config.variant.procurement_complexity),
        ("Variant self-build difficulty", config.variant.self_build_difficulty),
        ("Variant structural note", config.variant.structural_concept_note),
        ("Roof type", config.building.roof_type),
        ("Facade type", config.building.facade_type),
        ("Site", config.site.location_name),
        ("Climate zone", config.site.climate_zone),
        ("Altitude m", f"{config.site.altitude_m:g}"),
        ("Soil assumption", config.site.soil_assumption),
        ("Drainage risk", config.site.drainage_risk),
        ("Calculator warning", config.calculator_assumptions.warning),
        ("Smart home", f"{config.smart_home.platform} + {', '.join(config.smart_home.protocols)}"),
        ("Off-grid PV kWp", f"{config.off_grid.pv_kwp:g}"),
        ("Off-grid battery kWh", f"{config.off_grid.battery_kwh:g}"),
        ("Generator", config.off_grid.generator),
        ("Water tank L", f"{config.off_grid.water_tank_l:g}"),
        ("Wastewater options", "; ".join(config.off_grid.wastewater_options)),
        (
            "Supplier policy",
            "Use supplier placeholders only until real supplier quotations are provided.",
        ),
    )


def _build_assumptions_sheet(sheet: Worksheet, config: MountainRetreatConfig) -> None:
    sheet.append(("Assumption", "Value"))
    _style_header_row(sheet)
    for row in _assumption_rows(config):
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:B{sheet.max_row}"
    _set_widths(sheet, (34, 100))
    for cell in sheet["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet["B5"].font = NOTICE_FONT


def _apply_workbook_metadata(workbook: Workbook, config: MountainRetreatConfig) -> None:
    workbook.properties.title = "Mountain Retreat X1 Preliminary BOM"
    workbook.properties.subject = "Preliminary planning bill of materials"
    workbook.properties.creator = config.project.author
    workbook.properties.keywords = "PRELIMINARY, not for construction, Mountain Retreat X1"
    workbook.properties.created = DETERMINISTIC_WORKBOOK_TIMESTAMP
    workbook.properties.modified = DETERMINISTIC_WORKBOOK_TIMESTAMP


def generate_bom_workbook(
    config: MountainRetreatConfig,
    output_dir: Path,
    *,
    large_mode: bool = False,
) -> Path:
    """Generate the preliminary BOM workbook and return its path."""
    excel_dir = output_dir / BOM_OUTPUT_DIR
    excel_dir.mkdir(parents=True, exist_ok=True)

    materials = _all_materials(config, large_mode=large_mode)
    grouped = _materials_by_sheet(materials)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"

    for sheet_name in ITEM_SHEETS:
        workbook.create_sheet(sheet_name)
    assumptions = workbook.create_sheet("Assumptions")

    _build_summary_sheet(summary, materials)
    for sheet_name in ITEM_SHEETS:
        sheet = workbook[sheet_name]
        _configure_item_sheet(sheet)
        for item in grouped[sheet_name]:
            _append_material_row(sheet, item)
    _build_assumptions_sheet(assumptions, config)
    _apply_workbook_metadata(workbook, config)

    path = excel_dir / BOM_FILENAME
    workbook.save(path)
    return path


def _cost_items(config: MountainRetreatConfig) -> list[CostItem]:
    common_items = [
        item
        for item in config.cost_assumptions_serbia_2026.cost_items
        if item.code != "COST-002"
    ]
    return [*common_items, *config.variant.cost_items]


def _configure_cost_table_sheet(sheet: Worksheet) -> None:
    sheet.append(COST_ITEM_HEADERS)
    _style_header_row(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COST_ITEM_HEADERS))}1"
    _set_widths(sheet, (16, 18, 18, 42, 14, 10, 16, 18, 48))


def _append_cost_component_row(
    sheet: Worksheet,
    item: CostItem,
    unit_cost_attr: str,
) -> None:
    row = sheet.max_row + 1
    unit_cost = getattr(item, unit_cost_attr)
    sheet.append(
        (
            item.code,
            item.phase,
            item.category,
            item.description,
            item.quantity,
            item.unit,
            unit_cost,
            f"=E{row}*G{row}",
            item.notes,
        )
    )
    for cell in sheet[row]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in ("E", "G", "H"):
        sheet[f"{column}{row}"].number_format = EUR_FORMAT if column in ("G", "H") else "#,##0.00"


def _build_cost_component_sheet(
    sheet: Worksheet,
    items: Iterable[CostItem],
    unit_cost_attr: str,
) -> None:
    _configure_cost_table_sheet(sheet)
    for item in items:
        _append_cost_component_row(sheet, item, unit_cost_attr)
    total_row = sheet.max_row + 1
    sheet.append(("", "", "", "TOTAL", "", "", "", f"=SUM(H2:H{total_row - 1})", ""))
    for cell in sheet[total_row]:
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
    sheet[f"H{total_row}"].number_format = EUR_FORMAT


def _phase_or_category_formula(
    sheet_name: str,
    lookup_column: str,
    lookup_cell: str,
) -> str:
    return (
        f"SUMIF('{sheet_name}'!${lookup_column}:${lookup_column},"
        f"{lookup_cell},'{sheet_name}'!$H:$H)"
    )


def _build_cost_aggregation_sheet(
    sheet: Worksheet,
    labels: Iterable[str],
    lookup_column: str,
    title: str,
    contingency_rates: tuple[float, float, float],
) -> None:
    low_rate, standard_rate, high_rate = contingency_rates
    sheet.append(
        (
            title,
            "Material Subtotal",
            "Labor Subtotal",
            "Machinery Subtotal",
            "Subtotal Before Contingency",
            f"Contingency {low_rate:g}%",
            f"Contingency {standard_rate:g}%",
            f"Contingency {high_rate:g}%",
            f"Total With {high_rate:g}% Contingency",
            "Cost per Gross m2",
            "Cost per Net m2",
        )
    )
    _style_header_row(sheet)
    for row, label in enumerate(sorted(labels), start=2):
        lookup_cell = f"$A{row}"
        sheet.append(
            (
                label,
                f"={_phase_or_category_formula('Materials', lookup_column, lookup_cell)}",
                f"={_phase_or_category_formula('Labor', lookup_column, lookup_cell)}",
                f"={_phase_or_category_formula('Machinery', lookup_column, lookup_cell)}",
                f"=SUM(B{row}:D{row})",
                f"=E{row}*{low_rate:g}%",
                f"=E{row}*{standard_rate:g}%",
                f"=E{row}*{high_rate:g}%",
                f"=E{row}+H{row}",
                f"=I{row}/Assumptions!$B$11",
                f"=I{row}/Assumptions!$B$12",
            )
        )
    total_row = sheet.max_row + 1
    sheet.append(
        (
            "TOTAL",
            f"=SUM(B2:B{total_row - 1})",
            f"=SUM(C2:C{total_row - 1})",
            f"=SUM(D2:D{total_row - 1})",
            f"=SUM(E2:E{total_row - 1})",
            f"=SUM(F2:F{total_row - 1})",
            f"=SUM(G2:G{total_row - 1})",
            f"=SUM(H2:H{total_row - 1})",
            f"=SUM(I2:I{total_row - 1})",
            f"=I{total_row}/Assumptions!$B$11",
            f"=I{total_row}/Assumptions!$B$12",
        )
    )
    for cell in sheet[total_row]:
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=11):
        for cell in row:
            cell.number_format = EUR_FORMAT
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:K{sheet.max_row}"
    _set_widths(sheet, (22, 18, 16, 18, 24, 18, 18, 18, 24, 18, 18))


def _build_contingency_sheet(
    sheet: Worksheet,
    items: list[CostItem],
    contingency_rates: tuple[float, float, float],
) -> None:
    low_rate, standard_rate, high_rate = contingency_rates
    sheet.append(
        (
            "Cost Code",
            "Description",
            "Subtotal Before Contingency",
            "Item Contingency %",
            "Item Contingency",
            f"Contingency {low_rate:g}%",
            f"Contingency {standard_rate:g}%",
            f"Contingency {high_rate:g}%",
            "Total With Item Contingency",
        )
    )
    _style_header_row(sheet)
    for row, item in enumerate(items, start=2):
        source_row = row
        sheet.append(
            (
                item.code,
                item.description,
                f"=Materials!H{source_row}+Labor!H{source_row}+Machinery!H{source_row}",
                item.contingency_percent / 100,
                f"=C{row}*D{row}",
                f"=C{row}*{low_rate:g}%",
                f"=C{row}*{standard_rate:g}%",
                f"=C{row}*{high_rate:g}%",
                f"=C{row}+E{row}",
            )
        )
    total_row = sheet.max_row + 1
    sheet.append(
        (
            "TOTAL",
            "",
            f"=SUM(C2:C{total_row - 1})",
            "",
            f"=SUM(E2:E{total_row - 1})",
            f"=SUM(F2:F{total_row - 1})",
            f"=SUM(G2:G{total_row - 1})",
            f"=SUM(H2:H{total_row - 1})",
            f"=SUM(I2:I{total_row - 1})",
        )
    )
    for cell in sheet[total_row]:
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=9):
        for cell in row:
            cell.number_format = "0.00%" if cell.column == 4 else EUR_FORMAT
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:I{sheet.max_row}"
    _set_widths(sheet, (16, 42, 26, 18, 18, 18, 18, 18, 26))


def _build_scenario_sheet(
    sheet: Worksheet,
    items: list[CostItem],
    factor: float,
    contingency_percent: float,
) -> None:
    sheet.append(SCENARIO_HEADERS)
    _style_header_row(sheet)
    for row, item in enumerate(items, start=2):
        source_row = row
        sheet.append(
            (
                item.code,
                item.description,
                f"=Materials!H{source_row}+Labor!H{source_row}+Machinery!H{source_row}",
                factor,
                f"=C{row}*D{row}",
                contingency_percent,
                f"=E{row}*(1+F{row})",
                "Static placeholder scenario; contractor quotes required.",
            )
        )
    total_row = sheet.max_row + 1
    sheet.append(
        (
            "TOTAL",
            "",
            f"=SUM(C2:C{total_row - 1})",
            "",
            f"=SUM(E2:E{total_row - 1})",
            "",
            f"=SUM(G2:G{total_row - 1})",
            "",
        )
    )
    for cell in sheet[total_row]:
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=7):
        for cell in row:
            cell.number_format = "0.00%" if cell.column == 6 else EUR_FORMAT
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:H{sheet.max_row}"
    _set_widths(sheet, (16, 42, 30, 16, 18, 16, 24, 46))


def _build_off_grid_addon_sheet(sheet: Worksheet, config: MountainRetreatConfig) -> None:
    sheet.append(
        (
            "Add-on Code",
            "System",
            "Quantity Placeholder",
            "Unit",
            "Unit Cost Placeholder",
            "Subtotal",
            "Notes",
        )
    )
    _style_header_row(sheet)
    rows = config.cost_assumptions_serbia_2026.off_grid_addons
    for row, addon in enumerate(rows, start=2):
        system = str(addon["system"])
        assumption_ref = str(addon.get("assumption_ref", "cost_assumptions.off_grid_addons"))
        sheet.append(
            (
                str(addon["code"]),
                system,
                float(addon["quantity"]),
                str(addon["unit"]),
                float(addon["unit_cost_placeholder"]),
                f"=C{row}*E{row}",
                (
                    "Optional planning add-on only; professional design and quotes required. "
                    f"Assumption: {assumption_ref}."
                ),
            )
        )
    total_row = sheet.max_row + 1
    sheet.append(("TOTAL", "", "", "", "", f"=SUM(F2:F{total_row - 1})", ""))
    for cell in sheet[total_row]:
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = EUR_FORMAT
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{sheet.max_row}"
    _set_widths(sheet, (16, 42, 20, 12, 22, 18, 58))


def _build_cash_flow_sheet(sheet: Worksheet, phases: Iterable[str]) -> None:
    sheet.append(
        ("Phase", "Phase Total With Configured Contingency", "Cash Flow %", "Cash Flow EUR")
    )
    _style_header_row(sheet)
    phase_list = sorted(phases)
    total_row = len(phase_list) + 2
    for row, phase in enumerate(phase_list, start=2):
        sheet.append(
            (
                phase,
                f"=SUMIF(Cost_By_Phase!$A:$A,A{row},Cost_By_Phase!$I:$I)",
                f"=B{row}/SUM($B$2:$B${total_row - 1})",
                f"=B{row}",
            )
        )
    sheet.append(
        (
            "TOTAL",
            f"=SUM(B2:B{total_row - 1})",
            f"=SUM(C2:C{total_row - 1})",
            f"=SUM(D2:D{total_row - 1})",
        )
    )
    for cell in sheet[total_row]:
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = "0.00%" if cell.column == 3 else EUR_FORMAT
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:D{sheet.max_row}"
    _set_widths(sheet, (22, 30, 16, 18))


def _build_cost_assumptions_sheet(sheet: Worksheet, config: MountainRetreatConfig) -> None:
    cost_config = config.cost_assumptions_serbia_2026
    rows = (
        ("Project", config.project.project_name),
        ("Status", config.project.status),
        ("Currency", cost_config.currency),
        ("Assumption year", str(cost_config.assumption_year)),
        ("last_updated", cost_config.last_updated.isoformat()),
        ("Source policy", cost_config.source_policy),
        ("VAT included", str(cost_config.vat_included)),
        ("Default contingency", f"{cost_config.contingency_percent:g}%"),
        ("Regional adjustment note", cost_config.regional_adjustment_note),
        ("Gross area m2", f"{config.building.gross_area_m2:g}"),
        ("Net area m2", f"{config.building.net_area_m2:g}"),
        ("Estimate warning", cost_config.estimate_warning),
        ("Construction variant", config.variant.code),
        ("Active variant name", config.variant.name),
        ("Variant procurement complexity", config.variant.procurement_complexity),
        ("Off-grid PV kWp", f"{config.off_grid.pv_kwp:g}"),
        ("Off-grid battery kWh", f"{config.off_grid.battery_kwh:g}"),
        ("Professional limit", config.project.disclaimer),
        ("Quote requirement", "Contractor quotes are required before procurement or budgeting."),
    )
    sheet.append(("Assumption", "Value"))
    _style_header_row(sheet)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:B{sheet.max_row}"
    _set_widths(sheet, (30, 100))
    for cell in sheet["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _build_review_notes_sheet(sheet: Worksheet, config: MountainRetreatConfig) -> None:
    notes = (
        "All prices are placeholders and static planning assumptions.",
        "Prices are not current market quotes and must not be represented as supplier offers.",
        "Contractor quotes are required before procurement, financing, or construction decisions.",
        config.cost_assumptions_serbia_2026.source_policy,
        config.cost_assumptions_serbia_2026.regional_adjustment_note,
        config.cost_assumptions_serbia_2026.estimate_warning,
        (
            "Licensed professionals must review structural, electrical, mechanical, "
            "plumbing, and off-grid systems."
        ),
    )
    sheet.append(("Review Note",))
    _style_header_row(sheet)
    for note in notes:
        sheet.append((note,))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:A{sheet.max_row}"
    _set_widths(sheet, (120,))
    for cell in sheet["A"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _build_executive_summary(
    sheet: Worksheet,
    config: MountainRetreatConfig,
    item_count: int,
) -> None:
    sheet.append(("Metric", "Value", "Notes"))
    _style_header_row(sheet)
    cost_config = config.cost_assumptions_serbia_2026
    rates = _contingency_rates(cost_config.contingency_sensitivity_percents)
    component_total_row = item_count + 2
    scenario_total_row = item_count + 2
    off_grid_total_row = len(cost_config.off_grid_addons) + 2
    rows = (
        (
            "Material subtotal",
            f"=Materials!H{component_total_row}",
            "Formula references Materials total row once.",
        ),
        (
            "Labor subtotal",
            f"=Labor!H{component_total_row}",
            "Formula references Labor total row once.",
        ),
        (
            "Machinery subtotal",
            f"=Machinery!H{component_total_row}",
            "Formula references Machinery total row once.",
        ),
        ("Subtotal before contingency", "=SUM(B2:B4)", "Materials + labor + machinery."),
        (f"Contingency {rates[0]:g}%", f"=B5*{rates[0]:g}%", "YAML planning sensitivity."),
        (f"Contingency {rates[1]:g}%", f"=B5*{rates[1]:g}%", "YAML planning sensitivity."),
        (f"Contingency {rates[2]:g}%", f"=B5*{rates[2]:g}%", "YAML planning sensitivity."),
        (
            "Total with contingency",
            "=B5+B8",
            f"Uses configured high sensitivity of {rates[2]:g}%.",
        ),
        ("Cost per gross m2", "=B9/Assumptions!$B$11", "Gross area denominator."),
        ("Cost per net m2", "=B9/Assumptions!$B$12", "Net area denominator."),
        (
            "Optional off-grid add-on",
            f"=Scenario_OffGrid_AddOn!F{off_grid_total_row}",
            "Separate optional add-on from YAML.",
        ),
        ("Total with off-grid add-on", "=B9+B12", "Base total + optional add-on."),
        (
            "Economy scenario",
            f"=Scenario_Economy!G{scenario_total_row}",
            "Scenario comparison total.",
        ),
        (
            "Standard scenario",
            f"=Scenario_Standard!G{scenario_total_row}",
            "Scenario comparison total.",
        ),
        (
            "Premium scenario",
            f"=Scenario_Premium!G{scenario_total_row}",
            "Scenario comparison total.",
        ),
        (
            "last_updated",
            config.cost_assumptions_serbia_2026.last_updated.isoformat(),
            "From YAML cost assumptions.",
        ),
        ("Quote requirement", "Contractor quotes required", "Prices are placeholders."),
    )
    for row in rows:
        sheet.append(row)
    for row in sheet.iter_rows(min_row=2, max_row=16, min_col=2, max_col=2):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.number_format = EUR_FORMAT
    for row_number in (5, 8, 9, 13):
        for cell in sheet[row_number]:
            cell.font = BOLD_FONT
            cell.fill = SUBHEADER_FILL
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:C{sheet.max_row}"
    _set_widths(sheet, (34, 24, 58))


def _apply_cost_workbook_metadata(workbook: Workbook, config: MountainRetreatConfig) -> None:
    workbook.properties.title = "Mountain Retreat X1 Preliminary Cost Estimate"
    workbook.properties.subject = "Preliminary planning cost estimate"
    workbook.properties.creator = config.project.author
    workbook.properties.keywords = "PRELIMINARY, placeholder prices, contractor quotes required"
    workbook.properties.created = DETERMINISTIC_WORKBOOK_TIMESTAMP
    workbook.properties.modified = DETERMINISTIC_WORKBOOK_TIMESTAMP


def _contingency_rates(values: list[float]) -> tuple[float, float, float]:
    if len(values) >= 3:
        return (values[0], values[1], values[2])
    padded = [*values, 10, 15, 20]
    return (padded[0], padded[1], padded[2])


def generate_cost_estimate_workbook(config: MountainRetreatConfig, output_dir: Path) -> Path:
    """Generate the preliminary cost estimate workbook and return its path."""
    excel_dir = output_dir / COST_OUTPUT_DIR
    excel_dir.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.active.title = "Executive_Summary"
    for sheet_name in COST_WORKBOOK_SHEETS[1:]:
        workbook.create_sheet(sheet_name)

    items = _cost_items(config)
    phases = {item.phase for item in items}
    categories = {item.category for item in items}
    cost_config = config.cost_assumptions_serbia_2026
    contingency_rates = _contingency_rates(cost_config.contingency_sensitivity_percents)

    _build_cost_component_sheet(workbook["Materials"], items, "material_unit_cost")
    _build_cost_component_sheet(workbook["Labor"], items, "labor_unit_cost")
    _build_cost_component_sheet(workbook["Machinery"], items, "machinery_unit_cost")
    _build_cost_aggregation_sheet(
        workbook["Cost_By_Phase"],
        phases,
        "B",
        "Phase",
        contingency_rates,
    )
    _build_cost_aggregation_sheet(
        workbook["Cost_By_Category"],
        categories,
        "C",
        "Category",
        contingency_rates,
    )
    _build_contingency_sheet(workbook["Contingency"], items, contingency_rates)
    _build_cash_flow_sheet(workbook["Cash_Flow"], phases)
    _build_scenario_sheet(
        workbook["Scenario_Economy"],
        items,
        cost_config.scenario_factors.get("economy", 0.9),
        cost_config.scenario_contingency_percents.get("economy", 10) / 100,
    )
    _build_scenario_sheet(
        workbook["Scenario_Standard"],
        items,
        cost_config.scenario_factors.get("standard", 1.0),
        cost_config.scenario_contingency_percents.get("standard", 15) / 100,
    )
    _build_scenario_sheet(
        workbook["Scenario_Premium"],
        items,
        cost_config.scenario_factors.get("premium", 1.25),
        cost_config.scenario_contingency_percents.get("premium", 20) / 100,
    )
    _build_off_grid_addon_sheet(workbook["Scenario_OffGrid_AddOn"], config)
    _build_cost_assumptions_sheet(workbook["Assumptions"], config)
    _build_review_notes_sheet(workbook["Review_Notes"], config)
    _build_executive_summary(workbook["Executive_Summary"], config, len(items))
    _apply_cost_workbook_metadata(workbook, config)

    path = excel_dir / COST_FILENAME
    workbook.save(path)
    return path


def _default_gantt_tasks() -> tuple[GanttTask, ...]:
    return (
        GanttTask(
            "1.0",
            "Design and professional review",
            "Surveys, geotechnical brief, and licensed review",
            20,
            1,
            4,
            "",
            "Owner with licensed professionals",
            "Medium",
            "Yes",
            "Yes",
            "Mountain site access and weather windows must be confirmed.",
        ),
        GanttTask(
            "2.0",
            "Permits and site preparation",
            "Permitting pathway, temporary services, and site setup",
            15,
            5,
            7,
            "1.0",
            "Owner / architect / local authority",
            "High",
            "Yes",
            "Yes",
            "No generated document is a permit or approval.",
        ),
        GanttTask(
            "3.0",
            "Access road and logistics",
            "Access road verification, delivery planning, laydown zones",
            15,
            6,
            8,
            "1.0",
            "Contractor / civil consultant",
            "High",
            "Yes",
            "Yes",
            "Mountain logistics can control the real schedule.",
        ),
        GanttTask(
            "4.0",
            "Earthworks",
            "Excavation, slope control, and temporary erosion measures",
            15,
            8,
            10,
            "2.0, 3.0",
            "Civil contractor",
            "High",
            "Yes",
            "Yes",
            "Weather and soil assumptions must be verified on site.",
        ),
        GanttTask(
            "5.0",
            "Drainage",
            "Temporary and permanent drainage routes",
            10,
            9,
            11,
            "4.0",
            "Civil contractor / drainage designer",
            "High",
            "Yes",
            "Yes",
            "Drainage must be coordinated before foundations are closed.",
        ),
        GanttTask(
            "6.0",
            "Foundations",
            "Foundation works and slab concept execution",
            20,
            11,
            14,
            "4.0, 5.0",
            "Structural contractor",
            "High",
            "Yes",
            "Yes",
            "Requires geotechnical and structural design before construction.",
        ),
        GanttTask(
            "7.0",
            "Structural frame",
            "Timber/hybrid structural frame erection",
            20,
            15,
            18,
            "6.0",
            "Structural contractor",
            "High",
            "Yes",
            "Yes",
            "Lifting plan, bracing, and inspection hold points required.",
        ),
        GanttTask(
            "8.0",
            "Roof",
            "Roof structure, underlay, and standing seam concept",
            15,
            18,
            20,
            "7.0",
            "Roofing contractor",
            "High",
            "Yes",
            "Yes",
            "Snow, wind, fall protection, and weather closure are key risks.",
        ),
        GanttTask(
            "9.0",
            "Exterior closure",
            "Windows, doors, air/water barrier, temporary dry-in",
            15,
            20,
            22,
            "8.0",
            "Envelope contractor",
            "Medium",
            "Yes",
            "Yes",
            "Protect interior works from mountain weather.",
        ),
        GanttTask(
            "10.0",
            "Facade",
            "Timber, stone, and glass facade works",
            20,
            22,
            25,
            "9.0",
            "Facade contractor",
            "Medium",
            "Yes",
            "Yes",
            "Thermal bridges and waterproofing details require review.",
        ),
        GanttTask(
            "11.0",
            "Electrical rough-in",
            "Electrical containment, rough wiring, boards, exterior routes",
            15,
            24,
            26,
            "9.0",
            "Licensed electrician",
            "High",
            "Yes",
            "No",
            "Final cable sizing and code compliance by licensed professionals.",
        ),
        GanttTask(
            "12.0",
            "Plumbing rough-in",
            "Water, waste, vent, and drainage rough-in",
            15,
            24,
            26,
            "9.0",
            "Licensed plumber",
            "High",
            "Yes",
            "Yes",
            "Freeze protection and slopes must be inspected.",
        ),
        GanttTask(
            "13.0",
            "HVAC rough-in",
            "UFH manifolds, heat-pump interfaces, ventilation routes",
            15,
            25,
            27,
            "9.0, 11.0, 12.0",
            "Mechanical contractor",
            "High",
            "Yes",
            "Yes",
            "No final heat-loss calculation is generated by this project.",
        ),
        GanttTask(
            "14.0",
            "Interior walls and insulation",
            "Partitions, insulation, vapor control, service closures",
            20,
            27,
            30,
            "11.0, 12.0, 13.0",
            "Interior contractor",
            "Medium",
            "Yes",
            "No",
            "Do not close walls before MEP inspections.",
        ),
        GanttTask(
            "15.0",
            "Floors and finishes",
            "Floor build-ups, wall finishes, ceiling finishes",
            20,
            31,
            34,
            "14.0",
            "Finishing contractor",
            "Medium",
            "No",
            "No",
            "Moisture and temperature conditions affect finish quality.",
        ),
        GanttTask(
            "16.0",
            "Bathrooms",
            "Bathroom waterproofing, fixtures, extraction, finishes",
            15,
            32,
            34,
            "12.0, 14.0",
            "Bathroom contractor / plumber",
            "High",
            "Yes",
            "No",
            "Waterproofing hold points and ventilation checks required.",
        ),
        GanttTask(
            "17.0",
            "Kitchen",
            "Kitchen installation and appliance interfaces",
            10,
            35,
            36,
            "11.0, 12.0, 15.0",
            "Kitchen installer",
            "Medium",
            "No",
            "No",
            "Appliance loads and plumbing connections require review.",
        ),
        GanttTask(
            "18.0",
            "Terrace",
            "Terrace structure, drainage, guard, decking, utility placeholders",
            25,
            34,
            38,
            "7.0, 8.0, 10.0",
            "Terrace contractor / structural engineer",
            "High",
            "Yes",
            "Yes",
            "Terrace load, waterproofing, guard, and drainage risks are high.",
        ),
        GanttTask(
            "19.0",
            "Smart home",
            "Controller, network, sensors, cameras, scenes, tests",
            10,
            37,
            38,
            "11.0, 15.0",
            "Smart-home integrator",
            "Medium",
            "No",
            "No",
            "Security and life-safety integrations must remain advisory.",
        ),
        GanttTask(
            "20.0",
            "Solar/off-grid",
            "PV, battery, generator, water/off-grid placeholders",
            20,
            39,
            42,
            "8.0, 11.0, 13.0",
            "Licensed electrical/mechanical specialists",
            "High",
            "Yes",
            "Yes",
            "Winter performance and utility approvals are not guaranteed.",
        ),
        GanttTask(
            "21.0",
            "Landscaping",
            "Final grading, paths, erosion control, planting",
            20,
            43,
            46,
            "5.0, 18.0",
            "Landscape / civil contractor",
            "Medium",
            "Yes",
            "Yes",
            "Weather, slope, and drainage can shift sequencing.",
        ),
        GanttTask(
            "22.0",
            "Testing and handover",
            "Commissioning, inspections, owner handover, defect list",
            20,
            47,
            50,
            "16.0, 17.0, 18.0, 19.0, 20.0",
            "Contractor with licensed professionals",
            "High",
            "Yes",
            "Yes",
            "Contractor quotes, approvals, and commissioning records required.",
        ),
    )


def _configure_gantt_sheet(sheet: Worksheet) -> None:
    sheet.append(GANTT_HEADERS)
    _style_header_row(sheet)
    sheet.freeze_panes = "M2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(GANTT_HEADERS))}1"
    _set_widths(sheet, (10, 24, 42, 14, 12, 10, 24, 28, 12, 18, 18, 56, *([4] * 52)))
    for column in range(13, 65):
        sheet.cell(row=1, column=column).alignment = Alignment(
            text_rotation=90,
            horizontal="center",
            vertical="bottom",
        )


def _append_gantt_task(sheet: Worksheet, task: GanttTask) -> None:
    row = sheet.max_row + 1
    sheet.append(
        (
            task.wbs,
            task.phase,
            task.task,
            task.duration_days,
            task.start_week,
            task.end_week,
            task.dependencies,
            task.responsible_party,
            task.risk_level,
            task.inspection_required,
            task.weather_sensitive,
            task.notes,
            *("" for _ in range(52)),
        )
    )
    for cell in sheet[row]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    if task.risk_level == "High":
        sheet[f"I{row}"].fill = RISK_FILL
    for week in range(task.start_week, task.end_week + 1):
        cell = sheet.cell(row=row, column=12 + week)
        cell.value = "■"
        cell.fill = TIMELINE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_gantt_sheet(sheet: Worksheet, tasks: tuple[GanttTask, ...]) -> None:
    _configure_gantt_sheet(sheet)
    for task in tasks:
        _append_gantt_task(sheet, task)


def _build_phase_details_sheet(sheet: Worksheet, tasks: tuple[GanttTask, ...]) -> None:
    sheet.append(("WBS", "Phase", "Task", "Duration Days", "Window", "Responsible Party", "Notes"))
    _style_header_row(sheet)
    for task in tasks:
        sheet.append(
            (
                task.wbs,
                task.phase,
                task.task,
                task.duration_days,
                f"Week {task.start_week} to Week {task.end_week}",
                task.responsible_party,
                task.notes,
            )
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{sheet.max_row}"
    _set_widths(sheet, (10, 28, 44, 14, 20, 32, 70))


def _build_milestones_sheet(sheet: Worksheet, tasks: tuple[GanttTask, ...]) -> None:
    sheet.append(("Milestone", "Target Week", "Related WBS", "Notes"))
    _style_header_row(sheet)
    milestones = (
        ("Professional review complete", 4, "1.0", "Proceed only with required licensed input."),
        ("Site ready for earthworks", 8, "2.0, 3.0", "Access and logistics confirmed."),
        ("Foundations complete", 14, "6.0", "Inspection and hold-point records required."),
        ("Weather-tight shell", 22, "8.0, 9.0", "Dry-in before interior closures."),
        ("MEP rough-ins complete", 27, "11.0, 12.0, 13.0", "Do not close walls before review."),
        ("Terrace complete", 38, "18.0", "Guard, drainage, waterproofing checks required."),
        ("Off-grid systems tested", 42, "20.0", "No autonomy guarantee implied."),
        ("Handover complete", 50, "22.0", "Commissioning and owner documentation complete."),
    )
    for row in milestones:
        sheet.append(row)
        sheet.cell(row=sheet.max_row, column=2).fill = MILESTONE_FILL
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:D{sheet.max_row}"
    _set_widths(sheet, (34, 14, 18, 70))


def _build_dependencies_sheet(sheet: Worksheet, tasks: tuple[GanttTask, ...]) -> None:
    sheet.append(("WBS", "Phase", "Dependencies", "Dependency Warning"))
    _style_header_row(sheet)
    for task in tasks:
        warning = (
            "No predecessor listed; verify readiness before starting."
            if not task.dependencies
            else "Starting before predecessors close may create rework, inspection, or safety risk."
        )
        sheet.append((task.wbs, task.phase, task.dependencies or "None listed", warning))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:D{sheet.max_row}"
    _set_widths(sheet, (10, 30, 36, 78))


def _build_risks_sheet(sheet: Worksheet, tasks: tuple[GanttTask, ...]) -> None:
    sheet.append(("WBS", "Phase", "Risk Level", "Weather Sensitive", "Risk Notes"))
    _style_header_row(sheet)
    for task in tasks:
        notes = task.notes
        if task.weather_sensitive == "Yes":
            notes = f"{notes} Mountain weather can delay or resequence this phase."
        sheet.append((task.wbs, task.phase, task.risk_level, task.weather_sensitive, notes))
        if task.risk_level == "High":
            sheet.cell(row=sheet.max_row, column=3).fill = RISK_FILL
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:E{sheet.max_row}"
    _set_widths(sheet, (10, 30, 14, 18, 90))


def _build_inspections_sheet(
    sheet: Worksheet,
    config: MountainRetreatConfig,
    tasks: tuple[GanttTask, ...],
) -> None:
    sheet.append(("WBS/Checklist", "Phase", "Inspection Required", "Responsible Party", "Notes"))
    _style_header_row(sheet)
    for task in tasks:
        if task.inspection_required == "Yes":
            sheet.append((task.wbs, task.phase, "Yes", task.responsible_party, task.notes))
    for item in config.checklists_seed.checklist_items:
        sheet.append((item.id, item.phase, "Yes", item.responsible_party, item.inspection_item))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:E{sheet.max_row}"
    _set_widths(sheet, (18, 32, 20, 34, 86))


def _build_weekly_cashflow_sheet(sheet: Worksheet, tasks: tuple[GanttTask, ...]) -> None:
    sheet.append(("Week", "Planned Activity Count", "Planning Cashflow Weight", "Notes"))
    _style_header_row(sheet)
    for week in range(1, 53):
        active_count = sum(1 for task in tasks if task.start_week <= week <= task.end_week)
        sheet.append(
            (
                f"Week {week}",
                active_count,
                f"=B{week + 1}/SUM($B$2:$B$53)",
                "Placeholder distribution only; contractor cashflow required.",
            )
        )
    total_row = sheet.max_row + 1
    sheet.append(("TOTAL", f"=SUM(B2:B{total_row - 1})", f"=SUM(C2:C{total_row - 1})", ""))
    for cell in sheet[total_row]:
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
    for cell in sheet["C"]:
        cell.number_format = "0.00%"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:D{sheet.max_row}"
    _set_widths(sheet, (14, 24, 24, 70))


def _build_gantt_assumptions_sheet(sheet: Worksheet, config: MountainRetreatConfig) -> None:
    rows = (
        ("Project", config.project.project_name),
        ("Status", config.project.status),
        ("Disclaimer", config.project.disclaimer),
        ("Schedule duration", "12 months / 52 weeks"),
        ("Schedule basis", "Preliminary planning sequence only; not a contractor baseline."),
        ("Country", config.project.country),
        ("Climate zone", config.site.climate_zone),
        ("Altitude m", f"{config.site.altitude_m:g}"),
        ("Drainage risk", config.site.drainage_risk),
        ("Access road", config.site.access_road_type),
        ("Mountain weather note", "Snow, frost, rain, wind, and access limits can delay work."),
        (
            "Dependency warning",
            "Predecessor phases must be verified before starting dependent work.",
        ),
        ("Professional review", ", ".join(config.project.review_required_by)),
        ("YAML seed phases", str(len(config.construction_phases.phases))),
    )
    sheet.append(("Assumption", "Value"))
    _style_header_row(sheet)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:B{sheet.max_row}"
    _set_widths(sheet, (30, 100))
    for cell in sheet["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _apply_gantt_workbook_metadata(workbook: Workbook, config: MountainRetreatConfig) -> None:
    workbook.properties.title = "Mountain Retreat X1 Preliminary Gantt Schedule"
    workbook.properties.subject = "Preliminary 52-week planning schedule"
    workbook.properties.creator = config.project.author
    workbook.properties.keywords = "PRELIMINARY, Gantt, not a contractor baseline"
    workbook.properties.created = DETERMINISTIC_WORKBOOK_TIMESTAMP
    workbook.properties.modified = DETERMINISTIC_WORKBOOK_TIMESTAMP


def _gantt_tasks_from_config(config: MountainRetreatConfig) -> tuple[GanttTask, ...]:
    tasks: list[GanttTask] = []
    current_week = 1
    for phase in config.construction_phases.phases:
        duration_weeks = max(1, round(phase.duration_days / 5))
        start_week = min(current_week, 52)
        end_week = min(52, start_week + duration_weeks - 1)
        safety_text = " ".join(phase.safety_notes).lower()
        weather_sensitive = (
            "Yes"
            if any(token in safety_text for token in ("weather", "snow", "frost", "mountain"))
            else "No"
        )
        risk_level = "High" if phase.inspection_required or weather_sensitive == "Yes" else "Medium"
        notes = " ".join(phase.safety_notes) if phase.safety_notes else phase.description
        tasks.append(
            GanttTask(
                phase.wbs,
                phase.name,
                phase.description,
                phase.duration_days,
                start_week,
                end_week,
                ", ".join(phase.dependencies),
                phase.responsible_party,
                risk_level,
                "Yes" if phase.inspection_required else "No",
                weather_sensitive,
                notes,
            )
        )
        current_week = min(52, end_week + 1)
    return tuple(tasks)


def generate_gantt_schedule_workbook(config: MountainRetreatConfig, output_dir: Path) -> Path:
    """Generate the preliminary 52-week Gantt schedule workbook and return its path."""
    excel_dir = output_dir / GANTT_OUTPUT_DIR
    excel_dir.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.active.title = "Gantt"
    for sheet_name in GANTT_WORKBOOK_SHEETS[1:]:
        workbook.create_sheet(sheet_name)

    tasks = _gantt_tasks_from_config(config)
    _build_gantt_sheet(workbook["Gantt"], tasks)
    _build_phase_details_sheet(workbook["Phase_Details"], tasks)
    _build_milestones_sheet(workbook["Milestones"], tasks)
    _build_dependencies_sheet(workbook["Dependencies"], tasks)
    _build_risks_sheet(workbook["Risks"], tasks)
    _build_inspections_sheet(workbook["Inspections"], config, tasks)
    _build_weekly_cashflow_sheet(workbook["Cashflow_By_Week"], tasks)
    _build_gantt_assumptions_sheet(workbook["Assumptions"], config)
    _apply_gantt_workbook_metadata(workbook, config)

    path = excel_dir / GANTT_FILENAME
    workbook.save(path)
    return path


def _qa_phase_name(sheet_name: str) -> str:
    return sheet_name.replace("_", " ")


def _qa_scope_names(config: MountainRetreatConfig) -> tuple[str, ...]:
    room_names = tuple(
        room.name for room in [*config.rooms_ground_floor.rooms, *config.rooms_gallery.rooms]
    )
    terrace_names = tuple(f"Terrace - {zone.name}" for zone in config.terrace.zones)
    systems = (
        "Technical room",
        "Main distribution",
        "Water storage option",
        "Wastewater option",
        "Access road",
        "Foundation perimeter",
        "Roof drainage",
        "Owner handover set",
    )
    return (*room_names, *terrace_names, *systems)


def _qa_templates(sheet_name: str) -> tuple[tuple[str, str, str, str], ...]:
    raw_templates: dict[str, tuple[tuple[str, str, str], ...]] = {
        "Site_Preparation": (
            ("site", "Access route", "Owner / contractor"),
            ("site", "Setting out", "Surveyor"),
            ("safety", "Temporary works", "Safety lead"),
            ("environment", "Erosion control", "Civil contractor"),
            ("documentation", "Pre-start record", "Site manager"),
        ),
        "Earthworks": (
            ("civil", "Excavation depth", "Civil contractor"),
            ("geotechnical", "Soil condition", "Geotechnical reviewer"),
            ("safety", "Slope stability", "Safety lead"),
            ("civil", "Spoil management", "Site manager"),
            ("survey", "Formation level", "Surveyor"),
        ),
        "Drainage": (
            ("civil", "Drain route", "Drainage designer"),
            ("civil", "Filter layer", "Civil contractor"),
            ("waterproofing", "Interface", "Site manager"),
            ("maintenance", "Access", "Contractor"),
            ("weather", "Mountain runoff", "Civil reviewer"),
        ),
        "Foundations": (
            ("structural", "Subgrade approval", "Structural engineer"),
            ("structural", "Dimensions", "Site engineer"),
            ("waterproofing", "Damp proofing", "Waterproofing lead"),
            ("civil", "Frost protection", "Civil reviewer"),
            ("documentation", "Hold point", "Site manager"),
        ),
        "Rebar": (
            ("structural", "Bar placement", "Structural engineer"),
            ("structural", "Chairs/supports", "Site engineer"),
            ("structural", "Openings", "Structural engineer"),
            ("documentation", "Photo record", "Site manager"),
            ("safety", "Access", "Safety lead"),
        ),
        "Concrete": (
            ("concrete", "Delivery ticket", "Concrete contractor"),
            ("concrete", "Placement", "Site engineer"),
            ("concrete", "Curing", "Concrete contractor"),
            ("testing", "Samples", "Testing lab"),
            ("documentation", "Pour record", "Site manager"),
        ),
        "Timber_Structure": (
            ("structural", "Member condition", "Structural contractor"),
            ("structural", "Connections", "Structural engineer"),
            ("structural", "Bracing", "Site engineer"),
            ("envelope", "Moisture protection", "Contractor"),
            ("survey", "Plumb/level", "Surveyor"),
        ),
        "Roof": (
            ("roofing", "Underlay", "Roofing contractor"),
            ("roofing", "Metal covering", "Roofing contractor"),
            ("safety", "Fall protection", "Safety lead"),
            ("snow", "Snow measures", "Designer"),
            ("drainage", "Gutters/outlets", "Site manager"),
        ),
        "Windows_Doors": (
            ("envelope", "Opening prep", "Envelope contractor"),
            ("envelope", "Installation", "Site manager"),
            ("thermal", "Glazing", "Architect"),
            ("operation", "Function", "Installer"),
            ("documentation", "Warranty data", "Site manager"),
        ),
        "Facade": (
            ("envelope", "Substrate", "Facade contractor"),
            ("envelope", "Cladding", "Architect"),
            ("moisture", "Ventilation gap", "Site manager"),
            ("thermal", "Thermal bridge", "Envelope reviewer"),
            ("documentation", "Mockup/approval", "Architect"),
        ),
        "Electrical": (
            ("electrical", "Containment", "Licensed electrician"),
            ("electrical", "Box positions", "Licensed electrician"),
            ("electrical", "Labeling", "Electrical engineer"),
            ("safety", "Wet/exterior zones", "Electrical engineer"),
            ("testing", "Pre-close inspection", "Site manager"),
        ),
        "Plumbing": (
            ("plumbing", "Pipe routes", "Licensed plumber"),
            ("plumbing", "Waste/vent", "Plumbing designer"),
            ("plumbing", "Pressure test", "Licensed plumber"),
            ("waterproofing", "Penetrations", "Site manager"),
            ("documentation", "Fixture data", "Site manager"),
        ),
        "HVAC": (
            ("mechanical", "Heating loop", "Mechanical engineer"),
            ("mechanical", "Manifold/access", "HVAC contractor"),
            ("ventilation", "Extract/supply", "Mechanical engineer"),
            ("controls", "Thermostat", "HVAC contractor"),
            ("testing", "Commissioning prep", "HVAC contractor"),
        ),
        "Insulation": (
            ("thermal", "Continuity", "Envelope reviewer"),
            ("thermal", "Vapor control", "Site manager"),
            ("moisture", "Dryness", "Contractor"),
            ("fire", "Fire stopping", "Site manager"),
            ("documentation", "Photo record", "Site manager"),
        ),
        "Interior": (
            ("finishes", "Substrate", "Finishing contractor"),
            ("finishes", "Finish quality", "Architect"),
            ("joinery", "Fit-out", "Contractor"),
            ("safety", "Stairs/guards", "Architect"),
            ("handover", "Defects", "Site manager"),
        ),
        "Terrace": (
            ("terrace", "Structure", "Structural engineer"),
            ("waterproofing", "Membrane", "Waterproofing lead"),
            ("drainage", "Falls/drains", "Drainage designer"),
            ("safety", "Guard", "Architect"),
            ("finishes", "Decking/pavers", "Contractor"),
        ),
        "Smart_Home": (
            ("controls", "Device placement", "Smart-home integrator"),
            ("network", "Connectivity", "Network integrator"),
            ("security", "Privacy/safety", "Security reviewer"),
            ("testing", "Automation", "Smart-home integrator"),
            ("documentation", "Credentials", "Owner"),
        ),
        "Solar_OffGrid": (
            ("electrical", "PV/battery", "Electrical engineer"),
            ("electrical", "Protection", "Electrician"),
            ("mechanical", "Generator/fuel", "Specialist"),
            ("monitoring", "Energy data", "Integrator"),
            ("winter", "Snow/cold risk", "Owner"),
        ),
        "Final_Inspection": (
            ("handover", "Completion", "Site manager"),
            ("documentation", "Approvals", "Owner"),
            ("commissioning", "Systems", "Contractor"),
            ("safety", "Life safety", "Reviewer"),
            ("defects", "Punch list", "Site manager"),
        ),
        "Photo_Log": (
            ("records", "Photo before", "Site manager"),
            ("records", "Photo after", "Site manager"),
            ("records", "Defect photo", "Contractor"),
            ("records", "Weather photo", "Site manager"),
            ("records", "Closeout photo", "Owner"),
        ),
        "Document_Register": (
            ("records", "Drawing record", "Document controller"),
            ("records", "Submittal record", "Contractor"),
            ("records", "Inspection record", "Site manager"),
            ("records", "Warranty record", "Owner"),
            ("records", "Approval record", "Owner"),
        ),
    }
    phase = _qa_phase_name(sheet_name).lower()
    return tuple(
        (
            discipline,
            inspection,
            f"Confirm {inspection.lower()} for {{scope}} during {phase}.",
            responsible,
        )
        for discipline, inspection, responsible in raw_templates[sheet_name]
    )


def _configure_qa_sheet(sheet: Worksheet) -> None:
    sheet.append(QA_HEADERS)
    _style_header_row(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(QA_HEADERS))}1"
    _set_widths(sheet, (18, 24, 18, 44, 60, 28, 16, 18, 18, 14, 52, 24))


def _add_status_validation(sheet: Worksheet, end_row: int) -> None:
    quoted_values = ",".join(QA_STATUS_VALUES)
    validation = DataValidation(
        type="list",
        formula1=f'"{quoted_values}"',
        allow_blank=True,
    )
    validation.error = "Select a valid QA/QC status."
    validation.errorTitle = "Invalid status"
    validation.prompt = "Choose a status from the list."
    validation.promptTitle = "QA/QC status"
    sheet.add_data_validation(validation)
    validation.add(f"I2:I{end_row}")


def _qa_row_values(
    sheet_name: str,
    index: int,
    discipline: str,
    inspection: str,
    criteria: str,
    responsible: str,
) -> tuple[str, str, str, str, str, str, str, str, str, str, str, str]:
    phase = _qa_phase_name(sheet_name)
    prefix = "".join(part[0] for part in sheet_name.split("_"))
    required_photo = "Yes" if sheet_name not in {"Document_Register"} else "No"
    required_document = "Yes" if sheet_name not in {"Photo_Log"} else "No"
    return (
        f"QA-{prefix}-{index:04d}",
        phase,
        discipline,
        inspection,
        criteria,
        responsible,
        required_photo,
        required_document,
        "Not Started",
        "",
        "Preliminary QA/QC checklist row; verify against approved documents.",
        "",
    )


def _build_qa_sheet(
    sheet: Worksheet,
    config: MountainRetreatConfig,
    sheet_name: str,
    *,
    large_mode: bool = False,
) -> None:
    _configure_qa_sheet(sheet)
    templates = _qa_templates(sheet_name)
    scopes = _qa_scope_names(config)
    target_rows = 52 if large_mode else 16
    row_index = 1
    while row_index <= target_rows:
        scope = scopes[(row_index - 1) % len(scopes)]
        template = templates[(row_index - 1) % len(templates)]
        discipline, inspection, criteria, responsible = template
        sheet.append(
            _qa_row_values(
                sheet_name,
                row_index,
                discipline,
                f"{inspection} - {scope}",
                criteria.format(scope=scope),
                responsible,
            )
        )
        for cell in sheet[sheet.max_row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row_index += 1
    _add_status_validation(sheet, sheet.max_row)


def _apply_qa_workbook_metadata(workbook: Workbook, config: MountainRetreatConfig) -> None:
    workbook.properties.title = "Mountain Retreat X1 Preliminary QA/QC Checklists"
    workbook.properties.subject = "Preliminary construction QA/QC checklist workbook"
    workbook.properties.creator = config.project.author
    workbook.properties.keywords = "PRELIMINARY, QA/QC, not for construction"
    workbook.properties.created = DETERMINISTIC_WORKBOOK_TIMESTAMP
    workbook.properties.modified = DETERMINISTIC_WORKBOOK_TIMESTAMP


def _build_qa_assumptions_sheet(sheet: Worksheet, config: MountainRetreatConfig) -> None:
    sheet.append(("Assumption", "Value"))
    _style_header_row(sheet)
    rows = (
        ("Project", config.project.project_name),
        ("Status", config.project.status),
        ("Mandatory disclaimer", config.project.disclaimer),
        (
            "Professional review",
            ", ".join(config.project.review_required_by),
        ),
        (
            "Use limitation",
            (
                "QA/QC rows are preliminary planning prompts only and must be checked "
                "against approved construction documents, inspections, and local authority "
                "requirements."
            ),
        ),
        (
            "No approvals",
            (
                "This workbook does not create permits, approvals, signed inspections, "
                "or final acceptance."
            ),
        ),
    )
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:B{sheet.max_row}"
    _set_widths(sheet, (30, 110))
    for cell in sheet["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def generate_qa_checklist_workbook(
    config: MountainRetreatConfig,
    output_dir: Path,
    *,
    large_mode: bool = False,
) -> Path:
    """Generate the preliminary QA/QC checklist workbook and return its path."""
    excel_dir = output_dir / QA_OUTPUT_DIR
    excel_dir.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.active.title = QA_WORKBOOK_SHEETS[0]
    for sheet_name in QA_WORKBOOK_SHEETS[1:]:
        workbook.create_sheet(sheet_name)
    workbook.create_sheet(QA_ASSUMPTIONS_SHEET)

    for sheet_name in QA_WORKBOOK_SHEETS:
        _build_qa_sheet(workbook[sheet_name], config, sheet_name, large_mode=large_mode)
    _build_qa_assumptions_sheet(workbook[QA_ASSUMPTIONS_SHEET], config)
    _apply_qa_workbook_metadata(workbook, config)

    path = excel_dir / QA_FILENAME
    workbook.save(path)
    return path


def _maintenance_calendar_tasks(
    config: MountainRetreatConfig,
) -> tuple[MaintenanceCalendarTask, ...]:
    annual_years = tuple(range(1, 31))
    return (
        MaintenanceCalendarTask(
            "Monthly checklist",
            "Inspect visible leaks, terrace drains, technical room, bathrooms, and windows.",
            "Monthly",
            "Owner/operator",
            "1-2 hours",
            "Flashlight, camera, moisture meter if available",
            "Staining, swelling, musty smell, ponding, active drips",
            "Stop and call qualified trades if water ingress is suspected.",
            "Monthly",
            annual_years,
            "No, unless warning signs are found",
        ),
        MaintenanceCalendarTask(
            "Seasonal checklist",
            "Prepare cabin for winter freeze, snow, wind, and mountain access constraints.",
            "Autumn / before first frost",
            "Owner with HVAC/plumbing trades as needed",
            "0.5-1 day",
            "Drain-down checklist, insulation check, snow tools",
            "Unheated pipe routes, blocked drains, low frost protection",
            (
                f"Frost depth placeholder {config.site.frost_depth_placeholder_cm:g} "
                "cm is a risk reminder only."
            ),
            "Autumn",
            annual_years,
            "Yes, for plumbing/HVAC winterization details",
        ),
        MaintenanceCalendarTask(
            "Annual checklist",
            "Schedule professional roof, envelope, MEP, fireplace, and safety review.",
            "Annual",
            "Qualified contractor / licensed trades",
            "1 day allowance",
            "Access equipment, inspection reports, as-built notes",
            "Movement, leaks, corrosion, poor combustion, electrical faults",
            "Professional inspections do not create permits unless issued by authorities.",
            "Spring",
            annual_years,
            "Yes",
        ),
        MaintenanceCalendarTask(
            "5-year maintenance",
            (
                "Review coatings, sealants, membranes, exterior fixings, service "
                "history, and renewals."
            ),
            "Every 5 years",
            "Owner with architect/contractor",
            "1-3 days",
            "Inspection checklist, access equipment, supplier manuals",
            "UV damage, sealant cracking, timber finish breakdown, recurring leaks",
            "Use contractor quotes for renewal budgeting.",
            "Planning",
            (5, 10, 15, 20, 25, 30),
            "Yes",
        ),
        MaintenanceCalendarTask(
            "10-year maintenance",
            (
                "Review roof accessories, facade finish strategy, window hardware, "
                "HVAC major service needs."
            ),
            "Every 10 years",
            "Owner with specialist contractors",
            "2-5 days",
            "Inspection report, service records, replacement budget",
            "Corrosion, hardware wear, coating failure, lower HVAC efficiency",
            "Consider professional envelope and energy review before major spending.",
            "Planning",
            (10, 20, 30),
            "Yes",
        ),
        MaintenanceCalendarTask(
            "20-year maintenance",
            (
                "Commission major system renewal study for envelope, terrace, roof, "
                "HVAC, PV, and controls."
            ),
            "Year 20",
            "Owner with design team",
            "1-2 weeks planning",
            "Condition survey, budget model, energy review",
            "End-of-life equipment, repeated water ingress, obsolete controls",
            "Treat as a redesign checkpoint with licensed professional review.",
            "Planning",
            (20,),
            "Yes",
        ),
        MaintenanceCalendarTask(
            "30-year renewal planning",
            (
                "Prepare full renewal plan for envelope, structure interfaces, "
                "MEP systems, and interiors."
            ),
            "Year 30",
            "Owner with architect, engineers, and cost estimator",
            "2-4 weeks planning",
            "Condition survey, measured drawings, contractor quotes",
            "System obsolescence, hidden moisture, structural movement, code changes",
            "Do not assume original planning documents remain valid after 30 years.",
            "Planning",
            (30,),
            "Yes",
        ),
        MaintenanceCalendarTask(
            "Roof maintenance",
            (
                f"Inspect {config.building.roof_type}, gutters, snow guards, "
                "flashings, and penetrations."
            ),
            "Seasonal and after severe storms",
            "Owner for visual checks; roofer for access work",
            "2-4 hours",
            "Binoculars, camera, safe access equipment by qualified personnel",
            "Loose panels, corrosion, ice dams, blocked gutters, damaged flashing",
            "No roof access without fall protection and competent supervision.",
            "Spring/Autumn",
            annual_years,
            "Yes, for roof access or defects",
        ),
        MaintenanceCalendarTask(
            "Timber facade maintenance",
            (
                "Inspect timber cladding finish, ventilation gaps, end grain, "
                "fasteners, and splash zones."
            ),
            "Annual; refinish cycle TBD",
            "Owner / facade contractor",
            "0.5-1 day",
            "Moisture meter, camera, cleaning tools",
            "Cupping, rot, loose boards, trapped debris, coating breakdown",
            "Final coating cycles depend on selected product and exposure.",
            "Spring",
            annual_years,
            "Yes, if deterioration appears",
        ),
        MaintenanceCalendarTask(
            "Stone facade maintenance",
            "Inspect stone veneer/supports, joints, weeps, movement joints, and staining.",
            "Annual",
            "Masonry/facade contractor",
            "0.5 day",
            "Camera, joint probe by specialist, cleaning tools",
            "Cracked joints, displaced stone, efflorescence, blocked weeps",
            "Avoid aggressive cleaning without facade specialist review.",
            "Spring",
            annual_years,
            "Yes, if cracking or movement appears",
        ),
        MaintenanceCalendarTask(
            "Terrace maintenance",
            (
                "Inspect deck/paver surfaces, waterproofing terminations, drains, "
                "guards, stairs, and lighting."
            ),
            "Monthly in use season; seasonal deep check",
            "Owner / terrace contractor",
            "2-6 hours",
            "Hand tools, level, drain cleaning tools",
            "Ponding, loose guards, slippery surfaces, blocked drains, movement",
            f"Terrace area assumption is {config.terrace.terrace_area_m2:g} m2.",
            "Spring/Autumn",
            annual_years,
            "Yes, for guards, structure, waterproofing, or jacuzzi/fire-pit zones",
        ),
        MaintenanceCalendarTask(
            "Window and door maintenance",
            "Inspect glazing, seals, drainage slots, thresholds, hinges, locks, and weatherstrips.",
            "Seasonal",
            "Owner / window contractor",
            "2-4 hours",
            "Non-abrasive cleaner, hardware lubricant, camera",
            "Condensation between panes, air leakage, stiff operation, water at thresholds",
            "Large panoramic glazing requires specialist review if movement or cracking appears.",
            "Spring/Autumn",
            annual_years,
            "Yes, if glazing, structure, or egress concerns appear",
        ),
        MaintenanceCalendarTask(
            "HVAC maintenance",
            "Service heat pump, underfloor manifolds, filters, pumps, valves, and frost settings.",
            "Annual; filters 3-6 months",
            "Licensed mechanical/HVAC contractor",
            "0.5-1 day",
            "Manufacturer service kit, pressure/temperature readings",
            "Short cycling, low pressure, cold rooms, unusual noise, fault codes",
            "No final heat-loss calculation is included in this repository.",
            "Autumn",
            annual_years,
            "Yes",
        ),
        MaintenanceCalendarTask(
            "Fireplace maintenance",
            "Inspect fireplace/stove, flue, combustion air, hearth, clearances, and CO alarms.",
            "Before heating season",
            "Certified chimney/fireplace professional",
            "2-4 hours",
            "Chimney tools by professional, CO alarm tester",
            "Soot smell, poor draft, cracked glass, smoke spillage, CO alarm events",
            "Never operate after suspected flue or CO issue until inspected.",
            "Autumn",
            annual_years,
            "Yes",
        ),
        MaintenanceCalendarTask(
            "Solar PV maintenance",
            (
                "Inspect PV production, roof interfaces, isolators, labels, snow "
                "shading, and monitoring."
            ),
            "Quarterly visual; annual professional check",
            "Owner / licensed solar electrician",
            "1-3 hours",
            "Monitoring dashboard, camera, electrician test equipment",
            "Production drop, damaged cables, inverter faults, water at penetrations",
            f"PV assumption is {config.off_grid.pv_kwp:g} kWp and remains preliminary.",
            "Quarterly/Annual",
            annual_years,
            "Yes, for electrical inspection",
        ),
        MaintenanceCalendarTask(
            "Battery maintenance",
            (
                "Review battery area, BMS alarms, ventilation, temperature, "
                "clearances, and shutdown labels."
            ),
            "Monthly dashboard; annual professional check",
            "Owner / licensed electrical professional",
            "1-2 hours",
            "BMS dashboard, thermal scan by professional if needed",
            "Swelling, heat, odor, repeated faults, communication loss",
            f"Battery assumption is {config.off_grid.battery_kwh:g} kWh.",
            "Monthly/Annual",
            annual_years,
            "Yes, for battery area and BMS faults",
        ),
        MaintenanceCalendarTask(
            "Generator maintenance",
            (
                "Test generator start, fuel condition, exhaust route, transfer "
                "arrangement, and service interval."
            ),
            "Monthly exercise; annual service",
            "Owner / generator specialist / electrician",
            "1-3 hours",
            "Generator manual, fuel stabilizer, load test equipment by professional",
            "Hard starting, exhaust smell, fuel leaks, transfer faults, low runtime",
            f"Generator assumption: {config.off_grid.generator}. Transfer design is not finalized.",
            "Monthly/Annual",
            annual_years,
            "Yes, for transfer, exhaust, or fuel issues",
        ),
        MaintenanceCalendarTask(
            "Water system maintenance",
            (
                "Inspect tank, pump, filters, pressure, insulation, valves, leak "
                "sensors, and drain-down points."
            ),
            "Monthly; seasonal winterization",
            "Owner / licensed plumber",
            "1-4 hours",
            "Pressure gauge, filter cartridges, leak sensor test",
            "Pressure loss, cloudy water, pump cycling, leaks, frozen sections",
            f"Water tank option assumption: {config.off_grid.water_tank_l:g} L.",
            "Monthly/Autumn",
            annual_years,
            "Yes, for plumbing or water-quality concerns",
        ),
        MaintenanceCalendarTask(
            "Wastewater system maintenance",
            (
                "Review septic/biological treatment option, access covers, alarms, "
                "odors, and service contract."
            ),
            "Per authority/manufacturer; visual monthly",
            "Licensed wastewater service provider",
            "1-3 hours",
            "Service log, access tools by provider, alarm test",
            "Odor, backups, wet ground, alarm events, blocked venting",
            "Local wastewater approval and maintenance rules govern final operation.",
            "Monthly/Annual",
            annual_years,
            "Yes",
        ),
        MaintenanceCalendarTask(
            "Smart home maintenance",
            (
                "Back up Home Assistant, test Zigbee/Matter/MQTT devices, cameras, "
                "UPS, alerts, and credentials."
            ),
            "Monthly backup; annual security review",
            "Owner / smart-home integrator",
            "1-3 hours",
            "Admin dashboard, password manager, UPS test, spare batteries",
            "Offline devices, failed backups, stale updates, weak passwords, missing alerts",
            f"Configured platform assumption: {config.smart_home.platform}.",
            "Monthly/Annual",
            annual_years,
            "Yes, for security architecture changes",
        ),
        MaintenanceCalendarTask(
            "Emergency plan",
            (
                "Review emergency contacts, shutoff locations, generator procedure, "
                "freeze response, evacuation."
            ),
            "Annual and after major changes",
            "Owner/operator",
            "1-2 hours",
            "Printed emergency sheet, labels, flashlight, first-aid kit",
            "Unlabeled shutoffs, inaccessible equipment, expired fire extinguishers",
            "Keep paper instructions available because internet or power may fail.",
            "Annual",
            annual_years,
            "Yes, if life-safety systems change",
        ),
    )


def _configure_maintenance_sheet(sheet: Worksheet, headers: tuple[str, ...]) -> None:
    sheet.append(headers)
    _style_header_row(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    _set_widths(sheet, (10, 18, 26, 44, 22, 28, 18, 32, 40, 52, 28, 18, 28, 52))


def _build_maintenance_catalog(
    sheet: Worksheet, tasks: tuple[MaintenanceCalendarTask, ...]
) -> None:
    _configure_maintenance_sheet(
        sheet,
        ("Section", *MAINTENANCE_HEADERS, "Professional Review"),
    )
    for task in tasks:
        sheet.append(
            (
                task.section,
                task.task,
                task.frequency,
                task.responsible_person,
                task.estimated_effort,
                task.required_tools,
                task.warning_signs,
                task.notes,
                task.professional_review_required,
            )
        )
        for cell in sheet[sheet.max_row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _build_maintenance_calendar(
    sheet: Worksheet,
    tasks: tuple[MaintenanceCalendarTask, ...],
) -> None:
    _configure_maintenance_sheet(sheet, CALENDAR_HEADERS)
    for year in range(1, 31):
        for task in tasks:
            if year not in task.due_years:
                continue
            sheet.append(
                (
                    year,
                    task.month_season,
                    task.section,
                    task.task,
                    task.frequency,
                    task.responsible_person,
                    task.estimated_effort,
                    task.required_tools,
                    task.warning_signs,
                    task.notes,
                    task.professional_review_required,
                    "",
                    "",
                    "",
                )
            )
            for cell in sheet[sheet.max_row]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in range(2, sheet.max_row + 1):
        if sheet[f"A{row}"].value in {5, 10, 20, 30}:
            for cell in sheet[row]:
                cell.fill = SUBHEADER_FILL


def _build_year_plan(sheet: Worksheet) -> None:
    sheet.append(("Year", "Planning Focus", "Professional Review", "Notes"))
    _style_header_row(sheet)
    for year in range(1, 31):
        if year == 30:
            focus = "30-year renewal planning"
            review = "Architect, engineers, cost estimator, authority checks as required"
        elif year == 20:
            focus = "Major system renewal study"
            review = "Design team and specialist contractors"
        elif year % 10 == 0:
            focus = "10-year maintenance and medium renewal review"
            review = "Specialist contractors and envelope/MEP reviewers"
        elif year % 5 == 0:
            focus = "5-year maintenance and sealant/coating review"
            review = "Architect or contractor review"
        else:
            focus = "Routine monthly, seasonal, and annual maintenance"
            review = "Licensed trades where warning signs or regulated systems are involved"
        sheet.append(
            (
                year,
                focus,
                review,
                "PRELIMINARY planning aid; update from real service records.",
            )
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:D{sheet.max_row}"
    _set_widths(sheet, (10, 44, 54, 54))


def _build_professional_reviews(sheet: Worksheet) -> None:
    rows = (
        ("Roof/fall-risk work", "Roofer or competent contractor", "Before access work"),
        ("Structural movement/cracking", "Licensed structural engineer", "Immediately"),
        ("Electrical/PV/battery/generator", "Licensed electrical professional", "Before work"),
        ("HVAC/water/freeze protection", "Licensed mechanical/plumbing professional", "Annual"),
        ("Fireplace/flue/CO", "Certified chimney/fireplace professional", "Before heating season"),
        ("Wastewater", "Licensed wastewater provider/local authority", "Per permit/manufacturer"),
        ("Smart home/security", "Integrator/security reviewer", "Annual or after changes"),
    )
    sheet.append(("Topic", "Required Reviewer", "Timing"))
    _style_header_row(sheet)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:C{sheet.max_row}"
    _set_widths(sheet, (36, 52, 34))


def _build_emergency_plan(sheet: Worksheet) -> None:
    rows = (
        ("Water leak", "Shut off water, isolate power if safe, document leak, call plumber."),
        ("Electrical fault", "Do not reset repeatedly; isolate if safe and call electrician."),
        (
            "CO/fireplace alarm",
            "Evacuate, ventilate if safe, call emergency services/professional.",
        ),
        (
            "Freeze event",
            "Check frost-protection heat, water pressure, and drain-down procedure.",
        ),
        ("Generator fault", "Stop use if exhaust/fuel/electrical issue is suspected."),
        ("Battery alarm", "Follow manufacturer emergency instructions and call specialist."),
    )
    sheet.append(("Scenario", "Owner Response"))
    _style_header_row(sheet)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:B{sheet.max_row}"
    _set_widths(sheet, (30, 90))


def _build_maintenance_log(sheet: Worksheet) -> None:
    headers = (
        "Date",
        "Task",
        "Observation",
        "Responsible Person",
        "Photo/Document Ref",
        "Action Required",
        "Closeout Date",
    )
    sheet.append(headers)
    _style_header_row(sheet)
    for _ in range(40):
        sheet.append(("", "", "", "", "", "", ""))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{sheet.max_row}"
    _set_widths(sheet, (16, 34, 54, 28, 32, 48, 16))


def _apply_maintenance_workbook_metadata(
    workbook: Workbook,
    config: MountainRetreatConfig,
) -> None:
    workbook.properties.title = "Mountain Retreat X1 Maintenance Calendar"
    workbook.properties.subject = "Preliminary 30-year maintenance planning calendar"
    workbook.properties.creator = config.project.author
    workbook.properties.keywords = "PRELIMINARY, maintenance, not for construction"
    workbook.properties.created = DETERMINISTIC_WORKBOOK_TIMESTAMP
    workbook.properties.modified = DETERMINISTIC_WORKBOOK_TIMESTAMP


def generate_maintenance_calendar_workbook(
    config: MountainRetreatConfig,
    output_dir: Path,
) -> Path:
    """Generate the preliminary 30-year maintenance calendar workbook."""
    excel_dir = output_dir / MAINTENANCE_OUTPUT_DIR
    excel_dir.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.active.title = MAINTENANCE_WORKBOOK_SHEETS[0]
    for sheet_name in MAINTENANCE_WORKBOOK_SHEETS[1:]:
        workbook.create_sheet(sheet_name)

    tasks = _maintenance_calendar_tasks(config)
    _build_maintenance_calendar(workbook["Calendar"], tasks)
    _build_maintenance_catalog(workbook["Task_Catalog"], tasks)
    _build_year_plan(workbook["Year_1_to_30"])
    _build_professional_reviews(workbook["Professional_Reviews"])
    _build_emergency_plan(workbook["Emergency_Plan"])
    _build_maintenance_log(workbook["Maintenance_Log"])
    _build_assumptions_sheet(workbook["Assumptions"], config)
    _apply_maintenance_workbook_metadata(workbook, config)

    path = excel_dir / MAINTENANCE_FILENAME
    workbook.save(path)
    return path
